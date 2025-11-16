from __future__ import annotations

import asyncio
import time
from dataclasses import asdict, is_dataclass
from decimal import Decimal
from functools import partial
from io import BytesIO
from pathlib import Path
from typing import Dict, Any, List, Union, Optional, BinaryIO, Mapping

import pandas as pd
from openpyxl import load_workbook  # 只用来找表头行，懒得全换也可以不用

from tools.config_loader import load_config
from mydataclass.record import BidRecord

# =========================
# JSON 序列化与类型别名
# =========================

JsonScalar = Union[str, int, float, bool, None]
JsonValue = Union[JsonScalar, Dict[str, "JsonValue"], List["JsonValue"]]


def _json_default(o: Any):
    if isinstance(o, Decimal):
        return float(o)
    raise TypeError(f"Object of type {type(o).__name__} is not JSON serializable")


def _json_sanitize(v: Any) -> JsonValue:
    if isinstance(v, Decimal):
        return float(v)
    if is_dataclass(v):
        return _json_sanitize(asdict(v))
    if isinstance(v, (str, int, float, bool)) or v is None:
        return v
    if isinstance(v, dict):
        return {str(k): _json_sanitize(val) for k, val in v.items()}
    if isinstance(v, (list, tuple, set)):
        return [_json_sanitize(i) for i in v]
    return str(v)


# =========================
# Excel 表头映射（中文 -> 对外 JSON 字段名）
# =========================

HEADER_MAP_BASE: Dict[str, str] = {
    "内容": "content_text",        # 默认内容列
    "评级": "rating",
    "电话": "phone",
    "地区": "region",
    "羽色": "feather_color",
    "性别": "sex",                        # 或 "gender"
    "最高分速": "max_speed",
    "飞行距离": "distance",
    "资格赛": "qualifying_race",
    "第一关": "race_1",
    "第二关": "race_2",
    "双关鸽王": "double_race_king",
    "三关鸽王": "triple_race_king",
    "四关鸽王": "quad_race_king",
    "团体": "team",
    "本鸽收入": "pigeon_income",
    "鸽主总投入": "owner_total_invest",
    "鸽主总收入": "owner_total_income",
    "鸽主总营收": "owner_net_revenue",
    "开尔总营收": "kair_total_revenue",
    "缴费羽数": "paid_count",
    "进奖羽数": "award_count",
    "多羽入围": "multi_pigeon_in_award",
    "往年成绩": "past_results",
    "同年开尔他棚成绩": "same_year_kair_other_loft_result",
    "他棚成绩": "other_loft_result",
}


# =========================
# 全局 in-memory 数据：环号 -> 整行字段 dict
# =========================

RowsByRing = Dict[str, Dict[str, Any]]
XlsxSource = Union[str, Path, bytes, BinaryIO]

_ROWS_BY_RING: Optional[RowsByRing] = None
_ROWS_LOCK = asyncio.Lock()  # 防止并发初始化


# =========================
# 工具函数：环号规范化 / 表头定位
# =========================

def _normalize_ring(s: Optional[str]) -> str:
    if not s:
        return ""
    s = str(s).strip()
    s = (
        s.replace("－", "-")
         .replace("—", "-")
         .replace("–", "-")
         .replace("―", "-")
         .replace(" ", "")
    )
    return s.upper()


def _locate_header_row_with_openpyxl(
    xlsx_path: Path,
    ring_header: str = "环号",
    search_rows: int = 10,
) -> int:
    """
    可选：用 openpyxl 找一下“环号”那一行作为表头行。
    如果你 Excel 比较规整（第一行就是表头），这段逻辑其实可以省略。
    这里只是尽量兼容你之前“表头不一定在第一行”的情况。
    """
    wb = load_workbook(filename=str(xlsx_path), data_only=True, read_only=True)
    try:
        ws = wb.worksheets[0]
        max_r = min(search_rows, ws.max_row)
        for r in range(1, max_r + 1):
            values = [cell.value for cell in ws[r]]
            if ring_header in values:
                return r  # 认为这一行是表头行
    finally:
        wb.close()
    return 1  # 找不到就兜底用第一行


# =========================
# 同步函数：用 pandas 把 Excel 读成 rows_by_ring
# =========================

def _load_rows_by_ring_sync(
    xlsx: Optional[XlsxSource],
    *,
    ring_header: str,
    content_header: str,
    config_section: str,
    config_file: str,
) -> RowsByRing:
    # 1. 解析 xlsx 源
    if xlsx is None:
        cfg = load_config(section=config_section, file_path=config_file)
        if not isinstance(cfg, dict):
            raise RuntimeError("config section must be a dict")
        xlsx_path = cfg.get("xlsx_path") or cfg.get("xlsx")
        if not xlsx_path:
            raise RuntimeError("配置中未找到 xlsx_path / xlsx 字段")
        xlsx = Path(xlsx_path).expanduser().resolve()
    elif isinstance(xlsx, (str, Path)):
        xlsx = Path(xlsx).expanduser().resolve()

    # 2. 读 Excel，借助 openpyxl 找 header 行
    if isinstance(xlsx, Path):
        header_row_idx = _locate_header_row_with_openpyxl(
            xlsx,
            ring_header=ring_header,
            search_rows=10,
        )
        # pandas 的 header 参数：0-based
        df = pd.read_excel(str(xlsx), header=header_row_idx - 1, dtype=object)
    elif isinstance(xlsx, (bytes, bytearray)):
        df = pd.read_excel(BytesIO(xlsx), header=0, dtype=object)
    else:
        # BinaryIO
        df = pd.read_excel(xlsx, header=0, dtype=object)

    # 3. 确保有环号这一列
    if ring_header not in df.columns:
        raise RuntimeError(f"Excel 中未找到环号列: {ring_header}")

    # 4. 列名映射：中文 -> 英文
    header_map: Dict[str, str] = dict(HEADER_MAP_BASE)
    header_map[content_header] = "content_text"

    rename_map = {
        cn: en
        for cn, en in header_map.items()
        if cn in df.columns
    }
    df = df.rename(columns=rename_map)

    # 5. 把环号变成字符串 key
    df[ring_header] = df[ring_header].apply(
        lambda x: "" if pd.isna(x) else str(x).strip()
    )
    df = df[df[ring_header] != ""]

    # 6. 生成 rows_by_ring
    rows_by_ring: RowsByRing = {}
    for _, row in df.iterrows():
        ring = row[ring_header]
        row_dict: Dict[str, Any] = {}

        for col, val in row.items():
            if col == ring_header:
                continue
            if pd.isna(val):
                continue
            row_dict[col] = val

        if row_dict:
            rows_by_ring[ring] = row_dict

    return rows_by_ring


# =========================
# 异步初始化：在 on_startup 里调用一次
# =========================

async def init_pigeon_xlsx_context(
    *,
    xlsx: Optional[XlsxSource] = None,
    ring_header: str = "环号",
    content_header: str = "内容",
    config_section: str = "context",
    config_file: str = "config/spider.yaml",
) -> None:
    """
    异步友好的初始化：
    - 把重的 pandas 读 Excel 丢到线程池，不阻塞事件循环；
    - 填充全局 _ROWS_BY_RING；
    - 后续 records_to_payload 只读内存。
    """
    global _ROWS_BY_RING

    async with _ROWS_LOCK:
        if _ROWS_BY_RING is not None:
            return  # 已经初始化过了

        loop = asyncio.get_running_loop()
        fn = partial(
            _load_rows_by_ring_sync,
            xlsx,
            ring_header=ring_header,
            content_header=content_header,
            config_section=config_section,
            config_file=config_file,
        )
        rows = await loop.run_in_executor(None, fn)
        _ROWS_BY_RING = rows
        print(f"[INFO] pigeon xlsx context initialized by pandas, rows={len(rows)}")


# =========================
# 业务转换逻辑：单条 BidRecord
# =========================

def _one(r: BidRecord) -> Dict[str, Any]:
    user_code = r.user_code or ""
    results = _json_sanitize(getattr(r, "results", {}) or {})
    if not isinstance(results, dict):
        results = {}

    return {
        "id": getattr(r, "id", None),
        "code": getattr(r, "code", None),
        "auction_id": getattr(r, "auction_id", None),
        "pigeon_id": getattr(r, "pigeon_id", None),
        "pigeon_code": getattr(r, "pigeon_code", None),
        "pigeon_name": getattr(r, "pigeon_name", None),
        "auction_bid_count": getattr(r, "auction_bid_count", None),
        "auction_total_price": getattr(r, "auction_total_price", None),
        "auction_highest_price": getattr(r, "auction_highest_price", None),
        "auction_second_highest_price": getattr(r, "auction_second_highest_price", None),

        "auction_bid_count_all": getattr(r, "auction_bid_count_all", None),
        "auction_total_price_all": getattr(r, "auction_total_price_all", None),
        "auction_highest_price_all": getattr(r, "auction_highest_price_all", None),
        "auction_second_highest_price_all": getattr(r, "auction_second_highest_price_all", None),

        "match_score": getattr(r, "match_score", None),
        "user_id": getattr(r, "user_id", None),
        "user_code": user_code,
        "user_nickname": getattr(r, "user_nickname", None) or getattr(r, "usernickname", None),
        "user_avatar": getattr(r, "user_avatar", None),

        "type": getattr(r, "type", None),
        "margin": _json_sanitize(getattr(r, "margin", None)),
        "quote": _json_sanitize(getattr(r, "quote", None)),
        "count": _json_sanitize(getattr(r, "count", 0)),
        "status": getattr(r, "status", None),
        "status_time": getattr(r, "status_time", None),

        "create_user_id": getattr(r, "create_user_id", None),
        "create_admin_id": getattr(r, "create_admin_id", None),
        "create_time": getattr(r, "create_time", None),
        "cancel_user_id": getattr(r, "cancel_user_id", None),
        "cancel_admin_id": getattr(r, "cancel_admin_id", None),

        "results": results,
        "history": results.get(user_code, []),
    }


# =========================
# 用 rows_by_ring 丰富 current
# =========================

def _enrich_current_with_rows(
    current_info: dict,
    *,
    rows_by_ring: Mapping[str, Dict[str, Any]],
    target_field: str = "content",
) -> dict:
    enriched_current = dict(current_info or {})

    raw_footring = enriched_current.get("footring")
    footring = _normalize_ring(raw_footring)

    if not footring:
        print(f"[DEBUG] footring 为空: footring={raw_footring!r}")
        return enriched_current

    row = None

    # 原始环号匹配
    if raw_footring:
        raw_key = str(raw_footring).strip()
        row = rows_by_ring.get(raw_key)

    # 规范化匹配
    if row is None:
        normalized_view: Dict[str, Dict[str, Any]] = {}
        for k, v in rows_by_ring.items():
            if not k:
                continue
            nk = _normalize_ring(k)
            if nk:
                normalized_view[nk] = v
        row = normalized_view.get(footring)

    if row:
        enriched_current[target_field] = row
        print(f"[DEBUG] 匹配到行数据: {row}")
    else:
        print(f"[DEBUG] 未在 rows_by_ring 中找到足环: {footring}")

    return enriched_current


# =========================
# 装配最终 payload（外面调用保持不变）
# =========================

def records_to_payload(
    records: List[BidRecord],
    current_info: dict,
    *,
    target_field: str = "content",
) -> Dict[str, Any]:
    """
    外部调用：records_to_payload(records, current_info)

    前提：在程序启动阶段已经调用并 await 过 init_pigeon_xlsx_context()。
    """
    global _ROWS_BY_RING

    enriched_current = dict(current_info or {})

    rows_by_ring = _ROWS_BY_RING
    if rows_by_ring is not None:
        try:
            enriched_current = _enrich_current_with_rows(
                enriched_current,
                rows_by_ring=rows_by_ring,
                target_field=target_field,
            )
        except Exception as e:
            print(f"[DEBUG] rows_by_ring 匹配错误: {e}")
            enriched_current.setdefault("_xlsx_warning", str(e))
    else:
        print("[WARN] _ROWS_BY_RING 未初始化，records_to_payload 将不会附加 Excel 内容")

    return {
        "type": "pigeon/bids",
        "schema_version": "1.0",
        "ts": int(time.time() * 1000),
        "current_id": _json_sanitize(enriched_current),
        "items": [_one(r) for r in records],
    }


# =========================
# 错误 payload
# =========================

def error_payload(message: str, code: str = "INTERNAL") -> Dict[str, Any]:
    return {
        "type": "error",
        "schema_version": "1.0",
        "ts": int(time.time() * 1000),
        "error": {"code": code, "message": message},
    }
